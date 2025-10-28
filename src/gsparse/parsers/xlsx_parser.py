"""Parser for XLSX data."""

import logging
import io
from typing import List, Any, Dict, Optional
from openpyxl import load_workbook
from .base_parser import BaseParser
from ..core.spreadsheet import Spreadsheet
from ..core.worksheet import Worksheet

logger = logging.getLogger(__name__)


class XLSXParser(BaseParser):
	"""Parser for XLSX data from Google Sheets."""
	
	def __init__(self, data_only: bool = True):
		"""Initialize parser.
		
		Args:
			data_only: If True, only read cell values, not formulas
		"""
		self.data_only = data_only
	
	def parse(self, data: bytes, worksheet_name: str = "Sheet1") -> Worksheet:
		"""Parses XLSX data and returns Worksheet.
		
		Args:
			data: XLSX data bytes
			worksheet_name: Worksheet name
			
		Returns:
			Worksheet object
		"""
		try:
			# Load workbook from bytes
			workbook = load_workbook(
				io.BytesIO(data),
				data_only=self.data_only,
				read_only=True
			)
			
			# Get the specified worksheet
			if worksheet_name in workbook.sheetnames:
				worksheet = workbook[worksheet_name]
			else:
				# If worksheet not found, use the first one
				worksheet = workbook.active
				worksheet_name = worksheet.title
			
			# Convert worksheet to our format
			rows = []
			for row in worksheet.iter_rows(values_only=True):
				# Clean cell values
				cleaned_row = [self._clean_cell_value(cell) for cell in row]
				rows.append(cleaned_row)
			
			# Determine dimensions
			row_count = len(rows)
			column_count = max(len(row) for row in rows) if rows else 0
			
			# Normalize rows (pad with empty values to same length)
			normalized_rows = []
			for row in rows:
				normalized_row = row + [None] * (column_count - len(row))
				normalized_rows.append(normalized_row)
			
			workbook.close()
			
			return Worksheet(
				name=worksheet_name,
				data=normalized_rows,
				row_count=row_count,
				column_count=column_count
			)
			
		except Exception as e:
			logger.error(f"Error parsing XLSX data: {e}")
			raise ValueError(f"Failed to parse XLSX data: {e}")
	
	def parse_multiple(self, data_dict: Dict[str, bytes]) -> Spreadsheet:
		"""Parses multiple XLSX files and returns Spreadsheet.
		
		Args:
			data_dict: Dictionary {worksheet_name: XLSX_data}
			
		Returns:
			Spreadsheet object
		"""
		worksheets = []
		
		for worksheet_name, data in data_dict.items():
			worksheet = self.parse(data, worksheet_name)
			worksheets.append(worksheet)
		
		# Use first worksheet name as spreadsheet title
		title = list(data_dict.keys())[0] if data_dict else "Untitled"
		
		return Spreadsheet(
			title=title,
			worksheets=worksheets
		)
	
	def parse_workbook(self, data: bytes) -> Spreadsheet:
		"""Parses entire XLSX workbook and returns Spreadsheet with all worksheets.
		
		Args:
			data: XLSX data bytes
			
		Returns:
			Spreadsheet object with all worksheets
		"""
		try:
			# Load workbook from bytes
			workbook = load_workbook(
				io.BytesIO(data),
				data_only=self.data_only,
				read_only=True
			)
			
			worksheets = []
			
			# Process each worksheet
			for sheet_name in workbook.sheetnames:
				worksheet = workbook[sheet_name]
				
				# Convert worksheet to our format
				rows = []
				for row in worksheet.iter_rows(values_only=True):
					# Clean cell values
					cleaned_row = [self._clean_cell_value(cell) for cell in row]
					rows.append(cleaned_row)
				
				# Determine dimensions
				row_count = len(rows)
				column_count = max(len(row) for row in rows) if rows else 0
				
				# Normalize rows (pad with empty values to same length)
				normalized_rows = []
				for row in rows:
					normalized_row = row + [None] * (column_count - len(row))
					normalized_rows.append(normalized_row)
				
				worksheets.append(Worksheet(
					name=sheet_name,
					data=normalized_rows,
					row_count=row_count,
					column_count=column_count
				))
			
			workbook.close()
			
			# Use first worksheet name as spreadsheet title
			title = worksheets[0].name if worksheets else "Untitled"
			
			return Spreadsheet(
				title=title,
				worksheets=worksheets
			)
			
		except Exception as e:
			logger.error(f"Error parsing XLSX workbook: {e}")
			raise ValueError(f"Failed to parse XLSX workbook: {e}")
	
	def get_worksheet_names(self, data: bytes) -> List[str]:
		"""Gets list of worksheet names from XLSX file.
		
		Args:
			data: XLSX data bytes
			
		Returns:
			List of worksheet names
		"""
		try:
			workbook = load_workbook(
				io.BytesIO(data),
				read_only=True
			)
			
			sheet_names = workbook.sheetnames
			workbook.close()
			
			return sheet_names
			
		except Exception as e:
			logger.error(f"Error getting worksheet names: {e}")
			return []
